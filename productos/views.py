
import pandas as pd
from django.shortcuts import render, redirect
from .models import Producto, UbicacionCarrusel, ProductoGeneral,PedidoTemporal,ResultadoReactivo
from .forms import ExcelUploadForm, ExcelUbicacionesForm
from io import BytesIO
from django.http import HttpResponse, FileResponse
from django.contrib import messages
import os
import shutil
from django.utils.timezone import now
from django.db.models import Count
from django.utils.safestring import mark_safe
import math
from openpyxl import Workbook, load_workbook
from collections import Counter, defaultdict
from django.db.models import Sum, F

def subir_excel(request): 
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            nombre_archivo = archivo.name.lower()

            # Leer archivo CSV o Excel
            try:
                if nombre_archivo.endswith('.csv'):
                    primera_linea = archivo.readline().decode('utf-8')
                    archivo.seek(0)
                    if primera_linea.strip().lower().startswith("sep="):
                        df = pd.read_csv(archivo, sep=',', skiprows=1)
                    else:
                        df = pd.read_csv(archivo, sep=',')
                else:
                    df = pd.read_excel(archivo, engine='openpyxl')
            except Exception as e:
                messages.error(request, f"❌ Error al leer el archivo: {e}")
                return redirect('subir_excel')

            # Mapeo externo → interno
            renombres = {
                'codarticulo': 'cliente_codigo',
                'disponible_rack (sum)': 'stock_total',
                'saldo_galys (sum)': 'stock_carrusel',
            }

            # Renombrar columnas robustamente (ignorar comillas, mayúsculas, espacios)
            columnas_renombradas = {}
            for col in df.columns:
                normalizada = col.lower().strip().replace('"', '').replace("'", '')
                if normalizada in renombres:
                    columnas_renombradas[col] = renombres[normalizada]

            df.rename(columns=columnas_renombradas, inplace=True)
            df.columns = df.columns.str.lower().str.strip().str.replace('"', '').str.replace("'", '')

            # Verificar existencia de cliente_codigo
            if 'cliente_codigo' not in df.columns:
                messages.error(request, "❌ El archivo debe tener la columna 'cliente_codigo'.")
                return redirect('subir_excel')

            df['cliente_codigo'] = df['cliente_codigo'].astype(str).str.strip().str.upper()

            nuevos = []
            actualizados = []
            advertencias = []

            for _, row in df.iterrows():
                cliente_codigo = str(row['cliente_codigo']).strip().upper()
                if not cliente_codigo:
                    continue

                datos = {}
                for columna in df.columns:
                    if (
                        columna != 'cliente_codigo' and 
                        columna in [f.name for f in Producto._meta.get_fields()] and 
                        not pd.isna(row[columna])
                    ):
                        datos[columna] = row[columna]

                try:
                    producto = Producto.objects.get(cliente_codigo=cliente_codigo)
                    for campo, valor in datos.items():
                        setattr(producto, campo, valor)
                    producto.save()
                    actualizados.append(cliente_codigo)

                except Producto.DoesNotExist:
                    nuevos.append(Producto(cliente_codigo=cliente_codigo, **datos))

                except Producto.MultipleObjectsReturned:
                    advertencias.append(cliente_codigo)
                    continue

            if nuevos:
                Producto.objects.bulk_create(nuevos)

            mensaje = "✅ Archivo procesado con éxito. "
            if nuevos:
                mensaje += f"Se crearon {len(nuevos)} productos nuevos. "
            if actualizados:
                mensaje += f"Se actualizaron {len(actualizados)} productos existentes. "
            if advertencias:
                mensaje += f"⚠️ {len(advertencias)} códigos duplicados no fueron procesados: {', '.join(advertencias[:5])}..."

            messages.success(request, mensaje)
            return redirect('subir_excel')
    else:
        form = ExcelUploadForm()

    return render(request, 'subir_excel.html', {'form': form})

def descargar_plantilla(request):
    columnas = [
        'cliente_codigo', 'stock_total', 'stock_carrusel', 'cliente', 'codigo', 'descripcion', 
        'cantidad_por_caja', 'promedio_venta', 'promedio_sobredimensionado',
        'cantidad_op', 'tipo_ubicacion', 'unidades_por_batea', 'cantidad_bateas',
        'cantidad_max_bateas', 'stock_max_carrusel', 'psicofarmaco'
    ]
    df = pd.DataFrame(columns=columnas)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_RepoGalys.xlsx'
    return response

def inicio(request):
    return render(request, 'inicio.html')

def crear_backup(request):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    origen = os.path.join(base_dir, 'db.sqlite3')
    timestamp = now().strftime("%Y%m%d_%H%M%S")
    nombre_backup = f"backup_{timestamp}.sqlite3"

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'guardar':
            backup_dir = os.path.join(base_dir, 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            destino = os.path.join(backup_dir, nombre_backup)
            shutil.copy(origen, destino)
            mensaje = f"✅ Copia guardada en el servidor: {nombre_backup}"
            return render(request, 'backup.html', {'mensaje': mensaje})

        elif accion == 'descargar':
            destino_temporal = os.path.join(base_dir, 'temp_backup', nombre_backup)
            os.makedirs(os.path.dirname(destino_temporal), exist_ok=True)
            shutil.copy(origen, destino_temporal)
            response = FileResponse(open(destino_temporal, 'rb'), as_attachment=True, filename=nombre_backup)
            return response

    return render(request, 'backup.html')

def editar_producto(request):
    producto = None
    valores = {}
    campos = [f.name for f in Producto._meta.fields if f.name != 'id']

    if request.method == 'POST':
        cliente_codigo = request.POST.get('cliente_codigo')
        if 'buscar' in request.POST:
            producto = Producto.objects.filter(cliente_codigo=cliente_codigo).first()
            if producto:
                valores = {campo: getattr(producto, campo) for campo in campos}
            else:
                messages.info(request, "🔍 No se encontró el producto. Podés cargarlo desde cero.")
                valores = {'cliente_codigo': cliente_codigo}
        elif 'guardar' in request.POST:
            datos = {campo: request.POST.get(campo) for campo in campos}
            producto, creado = Producto.objects.update_or_create(
                cliente_codigo=datos['cliente_codigo'],
                defaults=datos
            )
            mensaje = "✅ Producto actualizado correctamente." if not creado else "🆕 Producto creado exitosamente."
            messages.success(request, mensaje)
            return redirect('editar_producto')

    return render(request, 'editar_producto.html', {
        'campos': campos,
        'valores': valores
    })

def subir_ubicaciones(request):
    if request.method == 'POST':
        form = ExcelUbicacionesForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            df = pd.read_excel(archivo)
            df.columns = df.columns.str.strip().str.lower()

            # Reemplazo de NaN por valores seguros
            df = df.fillna({
                'stock': 0,
                'entrando': 0,
                'saliendo': 0,
                'denominación': '',
                'lote': '',
                'caducidad': '',
                'udc': '',
                'udsudc': 0,
                'última entrada': '',
                'anchura (mm)': 0,
                'profundidad (mm)': 0,
                'altura (mm)': 0,
                'zona': '',
                'equipo': '',
                'módulo': '',
                'nivel': '',
                'fondo': '',
                'artículo': '',
                'reservado art.': '',
                'reservado udc': '',
                'fija': '',
                'codbarras1': '',
                'udc reserva': '',
                'bloqueada': '',
                'impedir entrada': '',
                'impedir salida': '',
                'udp': '',
                'udsudp': 0,
                'tp. stock': '',
                'propietario': '',
            })

            registros = []
            for _, row in df.iterrows():
                registros.append(UbicacionCarrusel(
                    id_posicion=str(row.get('idposiciondetalle', '')).strip(),
                    ubicacion=str(row.get('ubicación', '')).strip(),
                    stock=int(row.get('stock', 0)),
                    entrando=int(row.get('entrando', 0)),
                    saliendo=int(row.get('saliendo', 0)),
                    denominacion=row.get('denominación', ''),
                    lote=row.get('lote', ''),
                    caducidad=row.get('caducidad', ''),
                    udc=row.get('udc', ''),
                    uds_udc=int(row.get('udsudc', 0)),
                    ultima_entrada=row.get('última entrada', ''),
                    anchura=float(row.get('anchura (mm)', 0)),
                    profundidad=float(row.get('profundidad (mm)', 0)),
                    altura=float(row.get('altura (mm)', 0)),
                    zona=row.get('zona', ''),
                    equipo=row.get('equipo', ''),
                    modulo=row.get('módulo', ''),
                    nivel=row.get('nivel', ''),
                    fondo=row.get('fondo', ''),
                    articulo=row.get('artículo', ''),
                    reservado_articulo=row.get('reservado art.', ''),
                    reservado_udc=row.get('reservado udc', ''),
                    fija=row.get('fija', ''),
                    cod_barras=row.get('codbarras1', ''),
                    udc_reserva=row.get('udc reserva', ''),
                    bloqueada=row.get('bloqueada', ''),
                    impedir_entrada=row.get('impedir entrada', ''),
                    impedir_salida=row.get('impedir salida', ''),
                    udp=row.get('udp', ''),
                    uds_udp=int(row.get('udsudp', 0)),
                    tipo_stock=row.get('tp. stock', ''),
                    propietario=row.get('propietario', ''),
                ))

            # Limpiamos la tabla antes de la carga
            UbicacionCarrusel.objects.all().delete()
            UbicacionCarrusel.objects.bulk_create(registros)

            messages.success(request, f"✅ {len(registros)} ubicaciones cargadas correctamente.")
            return redirect('subir_ubicaciones')
    else:
        form = ExcelUbicacionesForm()

    return render(request, 'subir_ubicaciones.html', {'form': form})

def analisis_ocupacion(request):
    datos = UbicacionCarrusel.objects.exclude(ubicacion__iendswith='i')

    # Clasificaciones en texto
    clasificaciones = {
        0: "Vacias",
        1: "Menor a 10%",
        2: "Entre 10% y 20%",
        3: "Entre 20% y 30%",
        4: "Entre 30% y 40%",
        5: "Entre 40% y 50%",
        6: "Entre 50% y 60%",
        7: "Entre 60% y 70%",
        8: "Entre 70% y 80%",
        9: "Entre 80% y 90%",
        10: "Entre 90% y 100%",
    }

    # Traducción de alturas
    altura_labels = {
        100: "SUELO",
        180: "UDC170",
        380: "UDC320"
    }

    # Inicializar conteo
    resultado = {altura: {clas: 0 for clas in clasificaciones.values()} for altura in altura_labels.values()}
    valores_grafico = {altura: [0] * len(clasificaciones) for altura in altura_labels.values()}

    for u in datos:
        try:
            porcentaje = ((u.stock or 0) / (u.uds_udc or 1)) * 100
        except ZeroDivisionError:
            porcentaje = 0

        if porcentaje > 100:
            messages.warning(request, f"⚠️ {u.ubicacion}: porcentaje mayor al 100% ({porcentaje:.2f}%)")
            continue

        if u.uds_udc == 0 or u.stock == 0:
            clase = 0
        else:
            clase = min(int(porcentaje // 10) + 1, 10)

        clas_texto = clasificaciones[clase]
        altura = altura_labels.get(int(u.altura), "Otra")

        resultado[altura][clas_texto] += 1
        valores_grafico[altura][clase] += 1

    etiquetas = list(clasificaciones.values())
    columnas = list(resultado.keys())

    # Preparar los datos para la tabla: [(etiqueta, [valores por columna])]
    filas = [(etiquetas[i], [valores_grafico[col][i] for col in columnas]) for i in range(len(etiquetas))]

    contexto = {
        'resultado': resultado,
        'columnas': columnas,
        'etiquetas': etiquetas,
        'datos_grafico': [[valores_grafico[col][i] for col in columnas] for i in range(len(etiquetas))],
        'filas': filas,
    }
    # Calcular promedios
    ocupaciones_general = []
    ocupaciones_altura = {
        "SUELO": [],
        "UDC170": [],
        "UDC320": [],
    }

    totales_por_altura = {
        "SUELO": 20,
        "UDC170": 336,
        "UDC320": 840,
    }

    for u in datos:
        if u.stock is not None and u.uds_udc:
            try:
                porcentaje = (u.stock / u.uds_udc) * 100
                if porcentaje <= 100:  # filtramos errores
                    ocupaciones_general.append(porcentaje)

                    altura = altura_labels.get(int(u.altura), None)
                    if altura in ocupaciones_altura:
                        ocupaciones_altura[altura].append(porcentaje)
            except ZeroDivisionError:
                continue

    # Ahora agregamos ceros donde faltan
    for altura in totales_por_altura:
        cantidad_actual = len(ocupaciones_altura[altura])
        faltantes = totales_por_altura[altura] - cantidad_actual
        if faltantes > 0:
            ocupaciones_altura[altura].extend([0] * faltantes)

    # Finalmente calculamos los promedios
    todas = ocupaciones_altura["SUELO"] + ocupaciones_altura["UDC170"] + ocupaciones_altura["UDC320"]

    promedios = {
        "General": round(sum(todas) / len(todas), 2),
        "SUELO": round(sum(ocupaciones_altura["SUELO"]) / len(ocupaciones_altura["SUELO"]), 2),
        "UDC170": round(sum(ocupaciones_altura["UDC170"]) / len(ocupaciones_altura["UDC170"]), 2),
        "UDC320": round(sum(ocupaciones_altura["UDC320"]) / len(ocupaciones_altura["UDC320"]), 2),
    }
    contexto['promedios'] = promedios
    return render(request, 'analisis_ocupacion.html', contexto)

def calcular_bateas_requeridas(stock_carrusel, cantidad_a_reponer, unidades_por_batea):
    total_esperado = stock_carrusel + cantidad_a_reponer
    bateas_totales = math.ceil(total_esperado / unidades_por_batea) if unidades_por_batea else 0
    bateas_actuales = math.ceil(stock_carrusel / unidades_por_batea) if unidades_por_batea else 0
    return max(bateas_totales - bateas_actuales, 0)

def armar_reposicion(request):
    productos = Producto.objects.exclude(cliente__isnull=True).exclude(cliente='')

    if request.method == 'POST':
        filtro_psico = request.POST.get('psicofarmaco')
        if filtro_psico == 'SI':
            productos = productos.filter(psicofarmaco='SI')
        elif filtro_psico == 'NO':
            productos = productos.exclude(psicofarmaco='SI')

        dias_seleccionados = [int(d) for d in request.POST.getlist('dias')]
        alturas_seleccionadas = request.POST.getlist('alturas')
        if alturas_seleccionadas:
            productos = productos.filter(tipo_ubicacion__in=alturas_seleccionadas)

        clientes_seleccionados = request.POST.getlist('clientes')
        if clientes_seleccionados:
            productos = productos.filter(cliente__in=clientes_seleccionados)

        metodo_ocupacion = request.POST.get('metodo_ocupacion', 'simple')
        ajustar_porcentaje = request.POST.get('ajustar_porcentaje') == 'on'
        filtro_minimo_cantidad = request.POST.get('filtro_minimo_cantidad')
        minimo_unidades = int(filtro_minimo_cantidad) if filtro_minimo_cantidad else 0

        porcentaje_minimo_raw = request.POST.get('min_ocupacion')
        if porcentaje_minimo_raw == 'otro':
            try:
                porcentaje_minimo = float(request.POST.get('min_ocupacion_otro', '0'))
            except:
                porcentaje_minimo = 0
        else:
            porcentaje_minimo = float(porcentaje_minimo_raw or 0)

        resultados = []
        batea_contador = {"SUELO": 0, "UDC170": 0, "UDC320": 0}

        for p in productos:
            try:
                unidades_dia = p.promedio_sobredimensionado / 5 if p.promedio_sobredimensionado else 0
                dias_stock = round(p.stock_carrusel / unidades_dia) if unidades_dia else 0
                if dias_stock not in dias_seleccionados:
                    continue

                # Calcular cantidad base
                if p.stock_total < p.stock_max_carrusel - p.stock_carrusel:
                    cantidad = p.stock_total
                else:
                    faltante = p.stock_max_carrusel - p.stock_carrusel
                    cantidad = ((-(-faltante // p.cantidad_por_caja)) * p.cantidad_por_caja) if p.cantidad_por_caja else 0

                if cantidad <= 0:
                    continue

                # Calcular % ocupación base
                porcentaje = calcular_ocupacion_con_ubicaciones(p, cantidad) if metodo_ocupacion == 'ubicaciones' else calcular_ocupacion_simple(p, cantidad)

                # --- 🔧 Ajustar si no cumple mínimo ---
                if ajustar_porcentaje and porcentaje < porcentaje_minimo:
                    mejor_cantidad = 0
                    mejor_porcentaje = porcentaje
                    paso = p.cantidad_por_caja if tipo_altura == "SUELO" and p.cantidad_por_caja else 10
                    cantidad_reducida = cantidad

                    while cantidad_reducida >= paso:
                        cantidad_reducida -= paso
                        nuevo_porcentaje = calcular_ocupacion_con_ubicaciones(p, cantidad_reducida) if metodo_ocupacion == 'ubicaciones' else calcular_ocupacion_simple(p, cantidad_reducida)

                        if nuevo_porcentaje >= porcentaje_minimo:
                            mejor_cantidad = cantidad_reducida
                            mejor_porcentaje = nuevo_porcentaje
                            break

                    if mejor_cantidad > 0:
                        cantidad = mejor_cantidad
                        porcentaje = mejor_porcentaje
                    else:
                        if tipo_altura == "SUELO" and p.cantidad_por_caja:
                            if cantidad >= p.cantidad_por_caja:
                                cantidad = (cantidad // p.cantidad_por_caja) * p.cantidad_por_caja
                                porcentaje = calcular_ocupacion_con_ubicaciones(p, cantidad) if metodo_ocupacion == 'ubicaciones' else calcular_ocupacion_simple(p, cantidad)
                        else:
                            if cantidad >= 10:
                                cantidad = (cantidad // 10) * 10
                                porcentaje = calcular_ocupacion_con_ubicaciones(p, cantidad) if metodo_ocupacion == 'ubicaciones' else calcular_ocupacion_simple(p, cantidad)

                if cantidad < minimo_unidades:
                    continue

                tipo_altura = p.tipo_ubicacion
                bateas_usadas = 0
                if p.unidades_por_batea:
                    restante = p.stock_max_carrusel - p.stock_carrusel
                    cantidad_final = min(cantidad, restante)
                    bateas_usadas = math.ceil(cantidad_final / p.unidades_por_batea)

                if bateas_usadas > 0 and tipo_altura in batea_contador:
                    batea_contador[tipo_altura] += bateas_usadas

                resultados.append((p.cliente, p.codigo, cantidad, round(porcentaje), p.stock_carrusel))
            except:
                continue

        promedio_ocupacion = round(sum([r[3] for r in resultados]) / len(resultados), 1) if resultados else 0

        if request.POST.get('accion') == 'descargar':
            wb = Workbook()
            ws = wb.active
            ws.append(['Cliente', 'Código', 'Cantidad a reponer'])
            for fila in resultados:
                ws.append(fila[:3])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="reposicion_galys.xlsx"'
            wb.save(response)
            return response

        dias_opciones = [0, 1, 2, 3, 4, 5]
        alturas_opciones = ["SUELO", "UDC170", "UDC320"]
        clientes_opciones = list(Producto.objects.exclude(cliente__isnull=True).exclude(cliente='').values_list('cliente', flat=True).distinct().order_by('cliente'))

        return render(request, 'armar_reposicion.html', {
            'resultados': resultados,
            'dias_opciones': dias_opciones,
            'alturas_opciones': alturas_opciones,
            'clientes_opciones': clientes_opciones,
            'dias_seleccionados': dias_seleccionados,
            'alturas_seleccionadas': alturas_seleccionadas,
            'clientes_seleccionados': clientes_seleccionados,
            'psicofarmaco': filtro_psico,
            'cantidad_productos': len(resultados),
            'cantidad_unidades': sum([r[2] for r in resultados]),
            'bateas_necesarias': batea_contador,
            'porcentaje_minimo': porcentaje_minimo,
            'min_ocupacion': porcentaje_minimo_raw,
            'min_ocupacion_otro': request.POST.get('min_ocupacion_otro', ''),
            'metodo_ocupacion': metodo_ocupacion,
            'promedio_ocupacion': promedio_ocupacion,
            'ajustar_porcentaje': ajustar_porcentaje,
            'filtro_minimo_cantidad': filtro_minimo_cantidad,
        })

    # GET
    dias_opciones = [0, 1, 2, 3, 4, 5]
    alturas_opciones = ["SUELO", "UDC170", "UDC320"]
    clientes_opciones = list(Producto.objects.exclude(cliente__isnull=True).exclude(cliente='').values_list('cliente', flat=True).distinct().order_by('cliente'))

    return render(request, 'armar_reposicion.html', {
        'dias_opciones': dias_opciones,
        'alturas_opciones': alturas_opciones,
        'clientes_opciones': clientes_opciones,
        'dias_seleccionados': [],
        'alturas_seleccionadas': [],
        'clientes_seleccionados': [],
        'psicofarmaco': 'TODOS',
        'porcentaje_minimo': 0,
        'min_ocupacion': '75',
        'min_ocupacion_otro': '',
        'metodo_ocupacion': 'simple',
        'ajustar_porcentaje': False,
        'filtro_minimo_cantidad': '',
    })

def calcular_ocupacion_simple(p, cantidad):
    """
    Método estimado: calcula el porcentaje de ocupación de la última batea
    estimando con stock + cantidad / unidades_por_batea.
    """
    if not p.unidades_por_batea:
        return 0

    total = p.stock_carrusel + cantidad
    resto = total % p.unidades_por_batea

    if resto == 0 and total > 0:
        return 100
    else:
        return round((resto / p.unidades_por_batea) * 100, 2)


def calcular_ocupacion_con_ubicaciones(p, cantidad, ubicaciones):
    """
    Método preciso: analiza las ubicaciones reales del producto
    y simula la carga real para calcular ocupación de la última batea.
    """
    articulo = f"{p.cliente}-{p.codigo}"
    ubicaciones_producto = [u for u in ubicaciones if u.articulo == articulo]

    if not ubicaciones_producto or not p.unidades_por_batea:
        return calcular_ocupacion_simple(p, cantidad)

    # Ordenar por stock descendente (opcional)
    ubicaciones_producto.sort(key=lambda u: u.stock, reverse=True)

    cantidad_restante = cantidad
    ultima_ocupacion = 0

    for u in ubicaciones_producto:
        espacio_libre = u.uds_udc - u.stock
        if espacio_libre <= 0:
            continue

        asignar = min(cantidad_restante, espacio_libre)
        cantidad_restante -= asignar

        ocupacion = (u.stock + asignar) / u.uds_udc * 100
        ultima_ocupacion = round(ocupacion, 2)

        if cantidad_restante <= 0:
            break

    # Si quedó cantidad sin asignar, va a una batea nueva vacía
    if cantidad_restante > 0:
        ocupacion_extra = (cantidad_restante / p.unidades_por_batea) * 100
        ultima_ocupacion = round(ocupacion_extra, 2)

    return ultima_ocupacion

def reposicion_reactiva(request):
    resultados = []
    codigos_faltantes = []
    total_unidades = 0
    bateas_necesarias = {"SUELO": 0, "UDC170": 0, "UDC320": 0}
    porcentaje_total = 0

    if request.method == 'POST' and 'archivo' in request.FILES:
        PedidoTemporal.objects.all().delete()
        archivo = request.FILES['archivo']
        from openpyxl import load_workbook
        wb = load_workbook(filename=archivo, data_only=True)
        sheet = wb.active

        for fila in sheet.iter_rows(min_row=2, values_only=True):
            if not fila or all(cell is None for cell in fila):
                continue

            try:
                ubicacion = str(fila[18]).strip().lower() if len(fila) > 18 and fila[18] else ''
                if ubicacion.startswith("galys") or ubicacion.startswith("cfr"):
                    continue

                cliente = str(fila[4]).strip() if fila[4] else ''
                codigo = str(fila[1]).strip() if fila[1] else ''
                cantidad = int(fila[2]) if fila[2] else 0
                lote = str(fila[19]).strip() if len(fila) > 19 and fila[19] else ''

                if cliente and codigo and cantidad > 0:
                    PedidoTemporal.objects.create(
                        cliente=cliente,
                        codigo=codigo,
                        cantidad=cantidad,
                        lote=lote
                    )

            except Exception as e:
                print("Error procesando fila:", fila, "->", e)
                continue

    pedidos = PedidoTemporal.objects.all()
    acumulador = defaultdict(int)

    for pedido in pedidos:
        producto = Producto.objects.filter(cliente=pedido.cliente, codigo=pedido.codigo).first()

        if not producto:
            general = ProductoGeneral.objects.filter(cliente=pedido.cliente, codigo=pedido.codigo, galys=True).first()
            if not general:
                codigos_faltantes.append((pedido.cliente, pedido.codigo, pedido.lote))
                continue
            cantidad_por_caja = general.cantidad_por_caja or 1
            tipo = "UDC320"
            uds_batea = 156
        else:
            cantidad_por_caja = producto.cantidad_por_caja or 1
            tipo = producto.tipo_ubicacion.title() if producto.tipo_ubicacion else "UDC320"
            uds_batea = producto.unidades_por_batea or 1

        unidades_sueltas = pedido.cantidad % cantidad_por_caja
        if unidades_sueltas <= 0:
            continue

        clave = (pedido.cliente, pedido.codigo, pedido.lote)
        acumulador[clave] += unidades_sueltas
        total_unidades += unidades_sueltas

        bateas = math.ceil(unidades_sueltas / uds_batea)
        if tipo in bateas_necesarias:
            bateas_necesarias[tipo] += bateas

        porcentaje_total += (unidades_sueltas % uds_batea) / uds_batea * 100

    resultados = [(cliente, codigo, lote, cantidad) for (cliente, codigo, lote), cantidad in acumulador.items()]
    porcentaje_estimado = round(porcentaje_total / len(resultados), 2) if resultados else 0

    # Crear archivo de reposicion
    wb_reposicion = Workbook()
    ws = wb_reposicion.active
    ws.append(['Cliente', 'Código', 'Lote', 'Cantidad'])
    for fila in resultados:
        ws.append(fila)
    output_reposicion = BytesIO()
    wb_reposicion.save(output_reposicion)
    filename_reactiva = f"reposicion_reactiva_{timestamp}.xlsx"
    filepath_reactiva = f"/mnt/data/{filename_reactiva}"
    with open(filepath_reactiva, 'wb') as f:
        f.write(output_reposicion.getvalue())

    request.session['reposicion_reactiva_excel'] = filename_reactiva
    # Crear archivo de códigos faltantes
    wb_faltantes = Workbook()
    ws_f = wb_faltantes.active
    ws_f.append(['Cliente', 'Código', 'Lote'])
    for fila in codigos_faltantes:
        ws_f.append(fila)
    output_faltantes = BytesIO()
    wb_faltantes.save(output_faltantes)
    request.session['codigos_faltantes_excel'] = output_faltantes.getvalue()

    return render(request, 'reposicion_reactiva.html', {
        'resultados': resultados,
        'cantidad_productos': len(resultados),
        'cantidad_unidades': total_unidades,
        'bateas_necesarias': bateas_necesarias,
        'porcentaje_estimado': porcentaje_estimado,
        'hay_faltantes': len(codigos_faltantes) > 0
    })


def descargar_reposicion_reactiva(request):
    contenido = request.session.get('reposicion_reactiva_excel')
    if contenido:
        response = HttpResponse(
            contenido,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reposicion_reactiva.xlsx"'
        return response
    return HttpResponse("No hay archivo para descargar.")


def descargar_codigos_faltantes(request):
    contenido = request.session.get('codigos_faltantes_excel')
    if contenido:
        response = HttpResponse(
            contenido,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="codigos_faltantes.xlsx"'
        return response
    return HttpResponse("No hay archivo para descargar.")


def cargar_productos_generales(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        df = pd.read_excel(archivo)

        for _, row in df.iterrows():
            cliente = str(row.get('ProCliCodigo')).strip() if row.get('ProCliCodigo') else ''
            codigo = str(row.get('ProCodigo')).strip() if row.get('ProCodigo') else ''
            cantidad_caja = row.get('ProPacking')

            # Conversión más flexible para Galys
            valor = str(row.get('ProGalys')).strip().lower() if row.get('ProGalys') else ''
            galys = valor in ['verdadero', 'true', 'sí', 'si']

            if cliente and codigo:
                ProductoGeneral.objects.update_or_create(
                    cliente=cliente,
                    codigo=codigo,
                    defaults={
                        'galys': galys,
                        'cantidad_por_caja': cantidad_caja if pd.notnull(cantidad_caja) else None,
                    }
                )

        messages.success(request, "Archivo cargado correctamente.")
        return redirect('inicio')

    return render(request, 'cargar_productos_generales.html')

def comparar_cantidades(request):
    resultados = []

    if request.GET:  # Solo procesar si se presiona "Mostrar"
        ocultar_ceros = request.GET.get('ocultar_ceros') == '1'
        solo_sobrestock = request.GET.get('solo_sobrestock') == '1'

        productos = Producto.objects.exclude(cliente_codigo__isnull=True).exclude(cliente_codigo='')

        for p in productos:
            articulo = (p.cliente_codigo or "").strip().upper()
            ubicaciones = UbicacionCarrusel.objects.all()

            total_stock = 0
            bateas_ocupadas = 0

            for u in ubicaciones:
                articulo_ubi = (u.articulo or "").strip().upper()
                if articulo_ubi == articulo:
                    total_stock += u.stock or 0
                    bateas_ocupadas += 1

            diferencia = (p.stock_carrusel or 0) - total_stock
            stock_max = p.stock_max_carrusel or 0
            max_bateas = p.cantidad_max_bateas or 0

            fila = {
                'articulo': p.cliente_codigo,
                'cantidad_producto': p.stock_carrusel or 0,
                'cantidad_ubicaciones': total_stock,
                'diferencia': diferencia,
                'stock_max': stock_max,
                'max_bateas': max_bateas,
                'bateas_ocupadas': bateas_ocupadas,
                'stock_excedido': total_stock > stock_max,
                'bateas_excedidas': bateas_ocupadas > max_bateas
            }

            resultados.append(fila)

        # Filtro: ocultar productos con todos los valores en cero
        if ocultar_ceros:
            resultados = [
                r for r in resultados
                if any([
                    r['cantidad_producto'],
                    r['cantidad_ubicaciones'],
                    r['diferencia'],
                    r['stock_max'],
                    r['max_bateas'],
                    r['bateas_ocupadas']
                ])
            ]

        # Filtro: solo mostrar productos con sobrestock
        if solo_sobrestock:
            resultados = [r for r in resultados if r['cantidad_ubicaciones'] > r['stock_max']]

        resultados = sorted(resultados, key=lambda x: x['diferencia'], reverse=True)

    return render(request, 'comparar_cantidades.html', {
        'resultados': resultados,
        'request': request
    })